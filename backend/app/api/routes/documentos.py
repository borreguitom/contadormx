"""
Rutas para gestión de documentos fiscales por cliente.
POST /api/documentos/{cliente_id}/upload          — sube uno o varios archivos
GET  /api/documentos/{cliente_id}                 — lista documentos del cliente
GET  /api/documentos/{cliente_id}/resumen         — resumen fiscal (IVA, ISR, totales)
GET  /api/documentos/{cliente_id}/exportar-excel  — descarga Excel con dos hojas
GET  /api/documentos/{cliente_id}/diot            — DIOT agrupada por RFC
DELETE /api/documentos/{doc_id}/documento         — elimina un documento
"""
from __future__ import annotations

import io
import re
import urllib.parse
from collections import defaultdict
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import AsyncSessionLocal, Cliente, Documento, get_db
from app.core.deps import get_current_user
from app.core.database import User
from app.services.doc_extractor import extract_document

router = APIRouter()

# carpeta base para almacenar archivos subidos (relativa al backend)
UPLOAD_DIR = Path(__file__).parent.parent.parent.parent / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
ALLOWED_EXTENSIONS = {".xml", ".pdf", ".jpg", ".jpeg", ".png", ".webp", ".tiff", ".tif"}

# ── Magic bytes por extensión ─────────────────────────────────────────────────
_MAGIC: dict[str, list[bytes]] = {
    ".pdf":  [b"%PDF"],
    ".jpg":  [b"\xff\xd8\xff"],
    ".jpeg": [b"\xff\xd8\xff"],
    ".png":  [b"\x89PNG"],
    ".webp": [b"RIFF"],
    ".tiff": [b"II*\x00", b"MM\x00*"],
    ".tif":  [b"II*\x00", b"MM\x00*"],
}


def _valid_magic(content: bytes, ext: str) -> bool:
    if ext == ".xml":
        head = content[:200].lstrip()
        return (
            head.startswith(b"<?xml")
            or head.startswith(b"<cfdi")
            or head.startswith(b"<Comprobante")
        )
    sigs = _MAGIC.get(ext, [])
    return not sigs or any(content[:16].startswith(s) for s in sigs)


def _cliente_path(user_id: int, cliente_id: int) -> Path:
    p = UPLOAD_DIR / str(user_id) / str(cliente_id)
    p.mkdir(parents=True, exist_ok=True)
    return p


async def _verify_cliente(
    cliente_id: int,
    current_user: User,
    db: AsyncSession,
) -> Cliente:
    result = await db.execute(
        select(Cliente).where(
            Cliente.id == cliente_id,
            Cliente.user_id == current_user.id,
        )
    )
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente


# ── Upload ────────────────────────────────────────────────────────────────────

@router.post("/{cliente_id}/upload")
async def upload_documentos(
    cliente_id: int,
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    cliente = await _verify_cliente(cliente_id, current_user, db)
    dest_dir = _cliente_path(current_user.id, cliente_id)

    results = []
    for upload in files:
        ext = Path(upload.filename or "").suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            results.append({"archivo": upload.filename, "estado": "error", "detalle": "Formato no soportado"})
            continue

        content = await upload.read()
        if len(content) > MAX_FILE_SIZE:
            results.append({"archivo": upload.filename, "estado": "error", "detalle": "Archivo muy grande (máx 10 MB)"})
            continue

        if not _valid_magic(content, ext):
            results.append({"archivo": upload.filename, "estado": "error", "detalle": "El contenido no corresponde al formato declarado"})
            continue

        # guarda el archivo
        safe_name = Path(upload.filename).name
        dest_path = dest_dir / safe_name
        # si ya existe, agrega sufijo
        counter = 1
        while dest_path.exists():
            dest_path = dest_dir / f"{Path(safe_name).stem}_{counter}{ext}"
            counter += 1

        dest_path.write_bytes(content)

        # extrae datos
        tipo = "xml" if ext == ".xml" else ("pdf" if ext == ".pdf" else "imagen")
        extracted = extract_document(upload.filename or "", content)

        doc = Documento(
            cliente_id=cliente_id,
            user_id=current_user.id,
            nombre_archivo=upload.filename or safe_name,
            tipo_archivo=tipo,
            file_path=str(dest_path),
            **{k: extracted[k] for k in (
                "uuid_cfdi", "tipo_comprobante", "serie", "folio", "fecha_emision",
                "emisor_rfc", "emisor_nombre", "receptor_rfc", "receptor_nombre",
                "subtotal", "descuento", "iva_trasladado", "iva_retenido", "isr_retenido",
                "total", "moneda", "tipo_cambio", "conceptos", "estado", "error_msg",
            )},
        )
        db.add(doc)
        await db.flush()
        results.append({
            "archivo": upload.filename,
            "estado": extracted["estado"],
            "doc_id": doc.id,
            "uuid_cfdi": extracted["uuid_cfdi"],
            "total": extracted["total"],
        })

    await db.commit()
    return {"procesados": len(results), "resultados": results}


# ── Lista ─────────────────────────────────────────────────────────────────────

@router.get("/{cliente_id}")
async def list_documentos(
    cliente_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _verify_cliente(cliente_id, current_user, db)

    result = await db.execute(
        select(Documento)
        .where(Documento.cliente_id == cliente_id, Documento.user_id == current_user.id)
        .order_by(Documento.created_at.desc())
    )
    docs = result.scalars().all()

    return [
        {
            "id": d.id,
            "nombre_archivo": d.nombre_archivo,
            "tipo_archivo": d.tipo_archivo,
            "uuid_cfdi": d.uuid_cfdi,
            "tipo_comprobante": d.tipo_comprobante,
            "fecha_emision": d.fecha_emision.isoformat() if d.fecha_emision else None,
            "emisor_rfc": d.emisor_rfc,
            "emisor_nombre": d.emisor_nombre,
            "receptor_rfc": d.receptor_rfc,
            "total": d.total,
            "iva_trasladado": d.iva_trasladado,
            "moneda": d.moneda,
            "estado": d.estado,
            "error_msg": d.error_msg,
            "created_at": d.created_at.isoformat(),
        }
        for d in docs
    ]


# ── Resumen fiscal ────────────────────────────────────────────────────────────

@router.get("/{cliente_id}/resumen")
async def resumen_fiscal(
    cliente_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _verify_cliente(cliente_id, current_user, db)

    result = await db.execute(
        select(Documento).where(
            Documento.cliente_id == cliente_id,
            Documento.user_id == current_user.id,
            Documento.estado == "extraido",
        )
    )
    docs = result.scalars().all()

    ingresos = [d for d in docs if d.tipo_comprobante == "I"]
    egresos = [d for d in docs if d.tipo_comprobante == "E"]

    def totalize(lst: list[Documento]) -> dict[str, float]:
        return {
            "subtotal": sum(d.subtotal or 0 for d in lst),
            "descuento": sum(d.descuento or 0 for d in lst),
            "iva_trasladado": sum(d.iva_trasladado or 0 for d in lst),
            "iva_retenido": sum(d.iva_retenido or 0 for d in lst),
            "isr_retenido": sum(d.isr_retenido or 0 for d in lst),
            "total": sum(d.total or 0 for d in lst),
            "cantidad": len(lst),
        }

    ing = totalize(ingresos)
    egr = totalize(egresos)

    iva_neto = ing["iva_trasladado"] - egr["iva_trasladado"]
    isr_cargo = ing["isr_retenido"] + egr["isr_retenido"]

    return {
        "total_documentos": len(docs),
        "ingresos": ing,
        "egresos": egr,
        "iva_neto_a_pagar": round(iva_neto, 2),
        "isr_retenido_total": round(isr_cargo, 2),
        "utilidad_bruta": round(ing["total"] - egr["total"], 2),
    }


# ── Exportar Excel ───────────────────────────────────────────────────────────

@router.get("/{cliente_id}/exportar-excel")
async def exportar_excel(
    cliente_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    cliente = await _verify_cliente(cliente_id, current_user, db)

    result = await db.execute(
        select(Documento)
        .where(Documento.cliente_id == cliente_id, Documento.user_id == current_user.id)
        .order_by(Documento.fecha_emision.desc())
    )
    docs = result.scalars().all()

    wb = Workbook()

    # ── Hoja 1: Facturas ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Facturas"

    HEADER_FILL = PatternFill("solid", fgColor="14532D")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="F0FDF4")

    headers = [
        "Archivo", "Tipo", "UUID CFDI", "Tipo Comprobante",
        "Fecha Emisión", "Emisor RFC", "Emisor Nombre",
        "Receptor RFC", "Receptor Nombre",
        "Subtotal", "Descuento", "IVA Trasladado", "IVA Retenido", "ISR Retenido",
        "Total", "Moneda", "Estado",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    TIPO_MAP = {"I": "Ingreso", "E": "Egreso", "T": "Traslado", "N": "Nómina", "P": "Pago"}

    for row_num, d in enumerate(docs, 2):
        fill = ALT_FILL if row_num % 2 == 0 else PatternFill()
        values = [
            d.nombre_archivo,
            d.tipo_archivo,
            d.uuid_cfdi or "",
            TIPO_MAP.get(d.tipo_comprobante or "", d.tipo_comprobante or ""),
            d.fecha_emision.strftime("%d/%m/%Y") if d.fecha_emision else "",
            d.emisor_rfc or "",
            d.emisor_nombre or "",
            d.receptor_rfc or "",
            d.receptor_nombre or "",
            d.subtotal or 0,
            d.descuento or 0,
            d.iva_trasladado or 0,
            d.iva_retenido or 0,
            d.isr_retenido or 0,
            d.total or 0,
            d.moneda or "MXN",
            d.estado,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            if col in (10, 11, 12, 13, 14, 15):  # columnas numéricas
                cell.number_format = '"$"#,##0.00'

    # Ancho automático
    col_widths = [30, 10, 38, 14, 14, 15, 30, 15, 30, 14, 12, 15, 13, 13, 14, 8, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # ── Hoja 2: Resumen Fiscal ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen Fiscal")

    def _sum(lst, field):
        return sum(getattr(d, field) or 0 for d in lst)

    ingresos = [d for d in docs if d.tipo_comprobante == "I" and d.estado == "extraido"]
    egresos  = [d for d in docs if d.tipo_comprobante == "E" and d.estado == "extraido"]

    ws2["A1"] = f"Resumen Fiscal — {cliente.razon_social}"
    ws2["A1"].font = Font(bold=True, size=14, color="14532D")
    ws2.merge_cells("A1:C1")

    rows_resumen = [
        ("", "", ""),
        ("CONCEPTO", "INGRESOS", "EGRESOS"),
        ("Facturas procesadas", len(ingresos), len(egresos)),
        ("Subtotal", _sum(ingresos, "subtotal"), _sum(egresos, "subtotal")),
        ("Descuentos", _sum(ingresos, "descuento"), _sum(egresos, "descuento")),
        ("IVA trasladado", _sum(ingresos, "iva_trasladado"), _sum(egresos, "iva_trasladado")),
        ("IVA retenido", _sum(ingresos, "iva_retenido"), _sum(egresos, "iva_retenido")),
        ("ISR retenido", _sum(ingresos, "isr_retenido"), _sum(egresos, "isr_retenido")),
        ("Total", _sum(ingresos, "total"), _sum(egresos, "total")),
        ("", "", ""),
        ("IVA neto a pagar", _sum(ingresos, "iva_trasladado") - _sum(egresos, "iva_trasladado"), ""),
        ("ISR retenido total", _sum(ingresos, "isr_retenido") + _sum(egresos, "isr_retenido"), ""),
        ("Utilidad bruta", _sum(ingresos, "total") - _sum(egresos, "total"), ""),
    ]

    for r_num, (label, ing_val, egr_val) in enumerate(rows_resumen, 2):
        ws2.cell(r_num, 1, label)
        ws2.cell(r_num, 2, ing_val)
        if egr_val != "":
            ws2.cell(r_num, 3, egr_val)

        if r_num == 3:  # header row
            for col in range(1, 4):
                c = ws2.cell(r_num, col)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = HEADER_FILL
        elif isinstance(ing_val, float):
            for col in (2, 3):
                ws2.cell(r_num, col).number_format = '"$"#,##0.00'

    for col, w in [(1, 28), (2, 18), (3, 18)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Serializar
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^\w\-]", "_", cliente.razon_social or "cliente")[:30]
    filename = f"ContadorMX_{safe_name}_facturas.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}"},
    )


# ── DIOT automática ───────────────────────────────────────────────────────────

@router.get("/{cliente_id}/diot")
async def generar_diot(
    cliente_id: int,
    formato: str = "json",   # "json" | "txt"
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Genera la DIOT (Declaración Informativa de Operaciones con Terceros)
    agrupando los egresos del cliente por RFC de proveedor.
    Art. 32-B LIVA — presentación mensual día 17.

    formato=json  → datos estructurados para el frontend
    formato=txt   → archivo .txt en formato SAT pipe-separated
    """
    await _verify_cliente(cliente_id, current_user, db)

    result = await db.execute(
        select(Documento).where(
            Documento.cliente_id == cliente_id,
            Documento.user_id == current_user.id,
            Documento.estado == "extraido",
            Documento.tipo_comprobante == "E",   # solo egresos
        )
    )
    egresos = result.scalars().all()

    # Agrupar por RFC emisor (el proveedor del egreso)
    proveedores: dict = defaultdict(lambda: {
        "rfc": "",
        "nombre": "",
        "tipo_tercero": "04",       # 04 = nacional
        "tipo_operacion": "85",     # 85 = otros (compras/gastos generales)
        "iva_16_pagado": 0.0,
        "iva_16_no_creditable": 0.0,
        "iva_retenido": 0.0,
        "isr_retenido": 0.0,
        "monto_operaciones": 0.0,
        "facturas": 0,
    })

    for doc in egresos:
        rfc = doc.emisor_rfc or "XAXX010101000"
        p = proveedores[rfc]
        p["rfc"] = rfc
        p["nombre"] = doc.emisor_nombre or ""
        p["iva_16_pagado"] = round(p["iva_16_pagado"] + (doc.iva_trasladado or 0), 2)
        p["iva_retenido"] = round(p["iva_retenido"] + (doc.iva_retenido or 0), 2)
        p["isr_retenido"] = round(p["isr_retenido"] + (doc.isr_retenido or 0), 2)
        p["monto_operaciones"] = round(p["monto_operaciones"] + (doc.subtotal or 0), 2)
        p["facturas"] += 1

    filas = sorted(proveedores.values(), key=lambda x: x["rfc"])

    if formato == "txt":
        # Formato SAT DIOT — campos pipe-separated
        # Columnas según layout oficial de la aplicación DIOT del SAT
        lineas = []
        for p in filas:
            lineas.append("|".join([
                p["tipo_tercero"],          # 1 tipo tercero
                p["tipo_operacion"],        # 2 tipo operación
                p["rfc"],                   # 3 RFC
                p["nombre"][:150],          # 4 nombre
                "",                         # 5 país residencia (vacío = nacional)
                "",                         # 6 nacionalidad
                "",                         # 7 número registro tributario extranjero
                str(round(p["monto_operaciones"], 2)),   # 8 actos gravados 16%
                "0",                        # 9 actos gravados 16% no creditable
                "0",                        # 10 actos gravados importaciones 16%
                "0",                        # 11 actos gravados importaciones 16% no creditable
                "0",                        # 12 actos gravados 0%
                "0",                        # 13 actos exentos
                "0",                        # 14 actos gravados importaciones exentas
                str(round(p["iva_16_pagado"], 2)),       # 15 IVA 16% pagado
                "0",                        # 16 IVA 16% no creditable
                "0",                        # 17 IVA importaciones pagado
                str(round(p["iva_retenido"], 2)),        # 18 IVA retenido por el contribuyente
                "0",                        # 19 IVA retenido por terceros
            ]))
        txt_content = "\n".join(lineas)
        buf = io.BytesIO(txt_content.encode("utf-8"))
        return StreamingResponse(
            buf,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="DIOT.txt"'},
        )

    # JSON para el frontend
    totales = {
        "proveedores": len(filas),
        "facturas": sum(p["facturas"] for p in filas),
        "monto_total_operaciones": round(sum(p["monto_operaciones"] for p in filas), 2),
        "iva_total_pagado": round(sum(p["iva_16_pagado"] for p in filas), 2),
        "iva_total_retenido": round(sum(p["iva_retenido"] for p in filas), 2),
        "isr_total_retenido": round(sum(p["isr_retenido"] for p in filas), 2),
    }

    return {
        "periodo": "Todos los egresos registrados",
        "fundamento": "Art. 32-B LIVA — presentación día 17 de cada mes",
        "totales": totales,
        "proveedores": filas,
    }


# ── Eliminar ──────────────────────────────────────────────────────────────────

@router.delete("/{doc_id}/documento")
async def delete_documento(
    doc_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Documento).where(
            Documento.id == doc_id,
            Documento.user_id == current_user.id,
        )
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    # elimina el archivo físico si existe
    if doc.file_path:
        try:
            Path(doc.file_path).unlink(missing_ok=True)
        except Exception:
            pass

    await db.delete(doc)
    await db.commit()
    return {"ok": True}
