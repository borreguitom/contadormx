"""
Módulo de empleados para ContadorMX.
Gestión de nómina, CRUD de empleados, importación/exportación Excel.
"""
import io
from datetime import date, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from app.core.database import get_db, Cliente, Empleado
from app.core.deps import get_current_user, User
from app.calculators.nomina import calcular_nomina

router = APIRouter()

# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class EmpleadoCreate(BaseModel):
    nombre_completo: str
    rfc: Optional[str] = None
    curp: Optional[str] = None
    nss: Optional[str] = None
    fecha_nacimiento: Optional[date] = None
    fecha_alta: date
    tipo_contrato: str = "indeterminado"
    periodicidad_pago: str = "quincenal"
    salario_diario: float
    departamento: Optional[str] = None
    puesto: Optional[str] = None
    banco: Optional[str] = None
    clabe: Optional[str] = None


class EmpleadoUpdate(BaseModel):
    nombre_completo: Optional[str] = None
    rfc: Optional[str] = None
    curp: Optional[str] = None
    nss: Optional[str] = None
    fecha_nacimiento: Optional[date] = None
    fecha_alta: Optional[date] = None
    tipo_contrato: Optional[str] = None
    periodicidad_pago: Optional[str] = None
    salario_diario: Optional[float] = None
    departamento: Optional[str] = None
    puesto: Optional[str] = None
    banco: Optional[str] = None
    clabe: Optional[str] = None


class EmpleadoResponse(BaseModel):
    id: int
    cliente_id: int
    nombre_completo: str
    rfc: Optional[str]
    curp: Optional[str]
    nss: Optional[str]
    fecha_nacimiento: Optional[date]
    fecha_alta: date
    fecha_baja: Optional[date]
    tipo_contrato: Optional[str]
    periodicidad_pago: Optional[str]
    salario_diario: float
    departamento: Optional[str]
    puesto: Optional[str]
    banco: Optional[str]
    clabe: Optional[str]
    is_active: bool

    class Config:
        from_attributes = True


class NominaRequest(BaseModel):
    periodo: str
    fecha_inicio: date
    fecha_fin: date
    otras_percepciones_global: float = 0.0
    vales_despensa_global: float = 0.0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_cliente_or_404(
    cliente_id: int,
    current_user: User,
    db: AsyncSession,
) -> Cliente:
    result = await db.execute(
        select(Cliente).where(
            Cliente.id == cliente_id,
            Cliente.user_id == current_user.id,
        )
    )
    cliente = result.scalar_one_or_none()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente


async def _get_empleado_or_404(
    emp_id: int,
    cliente_id: int,
    db: AsyncSession,
) -> Empleado:
    result = await db.execute(
        select(Empleado).where(
            Empleado.id == emp_id,
            Empleado.cliente_id == cliente_id,
        )
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")
    return emp


def _run_nomina_for_empleados(
    empleados: list,
    periodo: str,
    otras_percepciones_global: float,
    vales_despensa_global: float,
) -> tuple[list, dict]:
    """
    Ejecuta calcular_nomina para cada empleado activo y agrega totales.
    Devuelve (lista_resultados, totales).
    """
    resultados = []
    totales = {
        "total_percepciones": 0.0,
        "total_deducciones": 0.0,
        "total_neto": 0.0,
        "costo_total_empresa": 0.0,
        "total_isr": 0.0,
        "total_imss_obrero": 0.0,
        "total_imss_patronal": 0.0,
        "total_infonavit": 0.0,
    }

    for emp in empleados:
        salario_mensual = float(emp.salario_diario) * 30
        resultado = calcular_nomina(
            salario_mensual_bruto=salario_mensual,
            periodo=periodo,
            otras_percepciones=otras_percepciones_global,
            vales_despensa=vales_despensa_global,
        )

        percepciones = resultado["percepciones"]
        deducciones = resultado["deducciones"]
        costo_empresa = resultado["costo_empresa"]

        totales["total_percepciones"] += percepciones["total_percepciones"]
        totales["total_deducciones"] += deducciones["total_deducciones"]
        totales["total_neto"] += resultado["neto_a_pagar"]
        totales["costo_total_empresa"] += costo_empresa["costo_total_empresa"]
        totales["total_isr"] += deducciones["isr_retenido"]
        totales["total_imss_obrero"] += deducciones["imss_cuota_obrero"]
        totales["total_imss_patronal"] += costo_empresa["imss_cuota_patronal"]
        totales["total_infonavit"] += costo_empresa["infonavit_patron"]

        resultados.append({
            "id": emp.id,
            "nombre_completo": emp.nombre_completo,
            "rfc": emp.rfc,
            "nss": emp.nss,
            "puesto": emp.puesto,
            "departamento": emp.departamento,
            "salario_diario": float(emp.salario_diario),
            "salario_mensual": salario_mensual,
            "percepciones": percepciones,
            "deducciones": deducciones,
            "neto_a_pagar": resultado["neto_a_pagar"],
            "costo_empresa": costo_empresa,
        })

    # Round totals to 2 decimals
    for k in totales:
        totales[k] = round(totales[k], 2)

    return resultados, totales


def _build_nomina_xlsx(resultados: list, totales: dict, periodo: str, fecha_inicio: date, fecha_fin: date) -> io.BytesIO:
    """Construye el xlsx de nómina en memoria y devuelve el BytesIO."""
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Nómina ----
    ws = wb.active
    ws.title = "Nómina"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center")
    money_fmt = '#,##0.00'

    headers = [
        "#", "Nombre", "RFC", "NSS", "Puesto", "Depto",
        "Salario Diario", "Salario Mensual",
        "Total Percepciones", "ISR Retenido", "IMSS Obrero",
        "Total Deducciones", "Neto a Pagar", "Costo Empresa",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    money_cols = {7, 8, 9, 10, 11, 12, 13, 14}

    for row_idx, emp in enumerate(resultados, start=2):
        row_data = [
            row_idx - 1,
            emp["nombre_completo"],
            emp["rfc"] or "",
            emp["nss"] or "",
            emp["puesto"] or "",
            emp["departamento"] or "",
            emp["salario_diario"],
            emp["salario_mensual"],
            emp["percepciones"]["total_percepciones"],
            emp["deducciones"]["isr_retenido"],
            emp["deducciones"]["imss_cuota_obrero"],
            emp["deducciones"]["total_deducciones"],
            emp["neto_a_pagar"],
            emp["costo_empresa"]["costo_total_empresa"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in money_cols:
                cell.number_format = money_fmt

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    ws.column_dimensions["B"].width = 30  # Nombre más ancho
    ws.freeze_panes = "A2"

    # ---- Sheet 2: Totales ----
    ws2 = wb.create_sheet("Totales")

    ws2.append(["ContadorMX — Resumen de Nómina"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Periodo:", periodo])
    ws2.append(["Fecha inicio:", str(fecha_inicio)])
    ws2.append(["Fecha fin:", str(fecha_fin)])
    ws2.append(["Total empleados:", len(resultados)])
    ws2.append([])

    totales_rows = [
        ("Total Percepciones", totales["total_percepciones"]),
        ("Total Deducciones", totales["total_deducciones"]),
        ("Total Neto a Pagar", totales["total_neto"]),
        ("Costo Total Empresa", totales["costo_total_empresa"]),
        ("Total ISR Retenido", totales["total_isr"]),
        ("Total IMSS Obrero", totales["total_imss_obrero"]),
        ("Total IMSS Patronal", totales["total_imss_patronal"]),
        ("Total Infonavit", totales["total_infonavit"]),
    ]

    for label, value in totales_rows:
        row = ws2.max_row + 1
        label_cell = ws2.cell(row=row, column=1, value=label)
        value_cell = ws2.cell(row=row, column=2, value=value)
        label_cell.font = Font(bold=True)
        value_cell.number_format = money_fmt

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Excel template endpoint — must come before /{cliente_id} to avoid confusion
# ---------------------------------------------------------------------------

@router.get("/template")
async def get_template(
    current_user: User = Depends(get_current_user),
):
    """Descarga plantilla xlsx para importación masiva de empleados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"

    headers = [
        "nombre_completo", "rfc", "curp", "nss",
        "fecha_nacimiento", "fecha_alta",
        "tipo_contrato", "periodicidad_pago",
        "salario_diario", "departamento", "puesto",
        "banco", "clabe",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Two example rows
    example_rows = [
        [
            "Juan García López", "GALJ800101AB1", "GALJ800101HDFRRL09",
            "12345678901", "1980-01-01", "2024-01-15",
            "indeterminado", "quincenal", 500.00,
            "Contabilidad", "Contador", "BANAMEX", "002180701111111111",
        ],
        [
            "María Rodríguez Soto", "ROSM900215CD2", "ROSM900215MDFDRR04",
            "98765432101", "1990-02-15", "2024-03-01",
            "determinado", "mensual", 350.00,
            "Administración", "Asistente", "BBVA", "012345678901234567",
        ],
    ]

    for row_idx, row_data in enumerate(example_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "plantilla_empleados.xlsx"
    headers_response = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response,
    )


# ---------------------------------------------------------------------------
# CRUD endpoints
# ---------------------------------------------------------------------------

@router.get("/{cliente_id}", response_model=List[EmpleadoResponse])
async def list_empleados(
    cliente_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_cliente_or_404(cliente_id, current_user, db)

    result = await db.execute(
        select(Empleado)
        .where(Empleado.cliente_id == cliente_id, Empleado.is_active == True)
        .order_by(Empleado.nombre_completo)
    )
    return result.scalars().all()


@router.post("/{cliente_id}", response_model=EmpleadoResponse, status_code=201)
async def create_empleado(
    cliente_id: int,
    data: EmpleadoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_cliente_or_404(cliente_id, current_user, db)

    emp = Empleado(
        **data.model_dump(),
        cliente_id=cliente_id,
        user_id=current_user.id,
    )
    db.add(emp)
    await db.commit()
    await db.refresh(emp)
    return emp


@router.put("/{cliente_id}/{emp_id}", response_model=EmpleadoResponse)
async def update_empleado(
    cliente_id: int,
    emp_id: int,
    data: EmpleadoUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_cliente_or_404(cliente_id, current_user, db)
    emp = await _get_empleado_or_404(emp_id, cliente_id, db)

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(emp, field, value)

    await db.commit()
    await db.refresh(emp)
    return emp


@router.delete("/{cliente_id}/{emp_id}", status_code=200)
async def delete_empleado(
    cliente_id: int,
    emp_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft delete: registra fecha_baja y marca is_active=False."""
    await _get_cliente_or_404(cliente_id, current_user, db)
    emp = await _get_empleado_or_404(emp_id, cliente_id, db)

    emp.fecha_baja = date.today()
    emp.is_active = False
    await db.commit()
    return {"ok": True, "fecha_baja": str(emp.fecha_baja)}


# ---------------------------------------------------------------------------
# Excel import
# ---------------------------------------------------------------------------

@router.post("/{cliente_id}/import")
async def import_empleados(
    cliente_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Importación masiva desde xlsx.
    Columnas esperadas (fila 1 = encabezados):
    nombre_completo, rfc, curp, nss, fecha_nacimiento, fecha_alta,
    tipo_contrato, periodicidad_pago, salario_diario, departamento,
    puesto, banco, clabe
    """
    await _get_cliente_or_404(cliente_id, current_user, db)

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo abrir el archivo: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    # Map header names to column index (0-based)
    header_row = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_index = {name: idx for idx, name in enumerate(header_row)}

    def get_val(row, col_name, default=None):
        idx = col_index.get(col_name)
        if idx is None:
            return default
        raw = row[idx] if idx < len(row) else None
        return raw if raw not in ("", None) else default

    def parse_date(val) -> Optional[date]:
        if val is None:
            return None
        if isinstance(val, (date, datetime)):
            return val.date() if isinstance(val, datetime) else val
        try:
            return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
        except ValueError:
            return None

    importados = 0
    errores = []

    for row_num, row in enumerate(rows[1:], start=2):
        nombre = get_val(row, "nombre_completo")
        if not nombre or str(nombre).strip() == "":
            continue  # Skip empty rows

        try:
            fecha_alta_raw = get_val(row, "fecha_alta")
            fecha_alta = parse_date(fecha_alta_raw)
            if not fecha_alta:
                raise ValueError("fecha_alta inválida o vacía")

            salario_raw = get_val(row, "salario_diario", 0)
            try:
                salario_diario = float(salario_raw)
            except (TypeError, ValueError):
                raise ValueError(f"salario_diario inválido: {salario_raw!r}")

            emp = Empleado(
                cliente_id=cliente_id,
                user_id=current_user.id,
                nombre_completo=str(nombre).strip(),
                rfc=get_val(row, "rfc"),
                curp=get_val(row, "curp"),
                nss=get_val(row, "nss"),
                fecha_nacimiento=parse_date(get_val(row, "fecha_nacimiento")),
                fecha_alta=fecha_alta,
                tipo_contrato=get_val(row, "tipo_contrato", "indeterminado"),
                periodicidad_pago=get_val(row, "periodicidad_pago", "quincenal"),
                salario_diario=salario_diario,
                departamento=get_val(row, "departamento"),
                puesto=get_val(row, "puesto"),
                banco=get_val(row, "banco"),
                clabe=get_val(row, "clabe"),
            )
            db.add(emp)
            importados += 1
        except Exception as exc:
            errores.append({"fila": row_num, "error": str(exc)})

    if importados > 0:
        await db.commit()

    return {"importados": importados, "errores": errores}


# ---------------------------------------------------------------------------
# Bulk payroll run (JSON response)
# ---------------------------------------------------------------------------

@router.post("/{cliente_id}/nomina")
async def run_nomina(
    cliente_id: int,
    body: NominaRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_cliente_or_404(cliente_id, current_user, db)

    result = await db.execute(
        select(Empleado)
        .where(Empleado.cliente_id == cliente_id, Empleado.is_active == True)
        .order_by(Empleado.nombre_completo)
    )
    empleados = result.scalars().all()

    if not empleados:
        raise HTTPException(status_code=404, detail="No hay empleados activos para este cliente")

    resultados, totales = _run_nomina_for_empleados(
        empleados,
        body.periodo,
        body.otras_percepciones_global,
        body.vales_despensa_global,
    )

    return {
        "periodo": body.periodo,
        "fecha_inicio": str(body.fecha_inicio),
        "fecha_fin": str(body.fecha_fin),
        "total_empleados": len(resultados),
        "totales": totales,
        "empleados": resultados,
    }


# ---------------------------------------------------------------------------
# Payroll Excel export
# ---------------------------------------------------------------------------

@router.get("/{cliente_id}/nomina/excel")
async def export_nomina_excel(
    cliente_id: int,
    periodo: str = Query(..., description="semanal|catorcenal|quincenal|mensual"),
    fecha_inicio: date = Query(...),
    fecha_fin: date = Query(...),
    otras_percepciones_global: float = Query(0.0),
    vales_despensa_global: float = Query(0.0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_cliente_or_404(cliente_id, current_user, db)

    result = await db.execute(
        select(Empleado)
        .where(Empleado.cliente_id == cliente_id, Empleado.is_active == True)
        .order_by(Empleado.nombre_completo)
    )
    empleados = result.scalars().all()

    if not empleados:
        raise HTTPException(status_code=404, detail="No hay empleados activos para este cliente")

    resultados, totales = _run_nomina_for_empleados(
        empleados,
        periodo,
        otras_percepciones_global,
        vales_despensa_global,
    )

    buf = _build_nomina_xlsx(resultados, totales, periodo, fecha_inicio, fecha_fin)

    filename = f"nomina_{cliente_id}_{periodo}_{fecha_inicio}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
